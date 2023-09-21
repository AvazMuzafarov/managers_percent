# Аваз Музафаров, 21.09.2023
# 1) Какой тип сделок (новая/текущая) был преобладающим в октябре 2021?

# загрузка модуля openpyxl для работы с файлами Excel
import openpyxl

# создание книги и листа
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active

# задаю все необходимые переменные
month_start = 'Октябрь 2021'    # месяц, начало периода выборки
month_stop = 'Ноябрь 2021'   # месяц, конец периода
deal_type_data = {'новая':0,   # словарь типов и их количества за период
                  'текущая':0}


# собственно сам цикл, который будет перебирать все строки листа и проверять, не равна ли переменная month началу отсчета
for row_number in range(2, sheet.max_row+1):
    month = sheet.cell(row=row_number, column=3).value
    if month == month_start:     # если равна, то запускается следующий цикл
        for i in range(row_number+1, sheet.max_row+1):  # этот цикл проверяет статус сделки в выбрнном периоде (от month_start до month_stop)
            deal_type = sheet.cell(row=i, column=5).value
            if deal_type in deal_type_data:
                deal_type_data[deal_type] +=1

            if month == month_stop: # если цикл доходит до конца периода, то цикл обрывается
                break
    if month == month_stop: # конец периода выборки = обрыв цикла
        break

# нахождение преобладающего типа
type_more = max(list(deal_type_data.values()))
name_type_more = list(filter(lambda key: deal_type_data[key] == type_more, deal_type_data))
print("Преобладающий тип сделок за период " + month_start, name_type_more)

# сохраняю все полученные данные в книге data3.xlsx в той же директории
sheet['J2'] = "Преобладающий тип сделок за период " + month_start
sheet.column_dimensions['J'].width = 45
sheet['J3'] = str(name_type_more[:1])
book.save('data3.xlsx')

# Ответ 3: Смирнов лучший менеджер. ответ находится в ячейке J3 книги data2.xlsx.