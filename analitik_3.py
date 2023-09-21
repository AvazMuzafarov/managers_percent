# Аваз Музафаров, 21.09.2023
# 3) Кто из менеджеров привлек для компании больше всего денежных средств в сентябре 2021?

# загрузка модуля openpyxl для работы с файлами Excel
import openpyxl

# создание книги и листа
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active

# задаю все необходимые переменные
month_start = 'Сентябрь 2021'    # месяц, начало периода выборки
month_stop = 'Октябрь 2021'      # месяц, конец периода
status_paid = 'ОПЛАЧЕНО'     # статусы сделок, кторые войдут в выборку
managers = {'Иванов':0,      # словарь менеджеров и их показатели
            'Андреев':0,
            'Филимонова':0,
            'Смирнов':0,
            'Кузнецова':0,
            'Петрова':0,
            'Васильев':0,
            'Соколов':0,
            'Михайлов':0,
            'Попов':0}


# собственно сам цикл, который будет перебирать все строки листа и проверять, не равна ли переменная month началу отсчета
for row_number in range(2, sheet.max_row+1):
    month = sheet.cell(row=row_number, column=3).value
    if month == month_start:     # если равна, то запускается следующий цикл
        for i in range(row_number+1, sheet.max_row+1):  # этот цикл проверяет статус сделки в выбрнном периоде (от month_start до month_stop)
            status = sheet.cell(row=i, column=3).value
            sum = sheet.cell(row=i, column=2).value
            sale = sheet.cell(row=i, column=4).value
            if status == status_paid and sale in managers: # если статус сделки 'ОПЛАЧЕНО' и менеджер акиивный, то сумма этой сделки сохраняется в паказателе этого менеджера
                managers[sale] += sum

            if status == month_stop: # если цикл доходит до конца периода, то цикл обрывается
                break
    if month == month_stop: # конец периода выборки = обрыв цикла
        break

# нахождение лучшего менеджера по показателям за период
best_manager_score = max(list(managers.values()))
winner_manager = list(filter(lambda key: managers[key] == best_manager_score, managers))
print("Лучший менеджер за период " + month_start, winner_manager)     # если нужно узнать только менеджера, то в консоле можно просмотреть

# сохраняю все полученные данные в книге data2.xlsx в той же директории
sheet['J2'] = "Лучший менеджер за период " + month_start
sheet.column_dimensions['J'].width = 30
sheet['J3'] = str(winner_manager[:1])
book.save('data2.xlsx')

# Ответ 3: Смирнов лучший менеджер. ответ находится в ячейке J3 книги data2.xlsx.