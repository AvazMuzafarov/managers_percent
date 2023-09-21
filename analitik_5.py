# Аваз Музафаров, 21.09.2023
# 5) Сколько оригиналов договора по майским сделкам было получено в июне 2021?

# загрузка модуля
import openpyxl
import datetime

# создание книги и листа
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active

# задаю все необходимые переменные
month_start = 'Май 2021'    # месяц, начало периода выборки
month_stop = 'Июнь 2021'    # месяц, конец периода
document = 'оригинал'       # оригинал договора
document_amount = 0
ttt = datetime.datetime(2021, 5, 31, 1, 11, 11 )
rrr= datetime.datetime(2021, 7, 1, 1, 11, 11 )

# собственно сам цикл, который будет перебирать все строки листа и проверять, не равна ли переменная month началу отсчета
for row_number in range(2, sheet.max_row+1):
    month = sheet.cell(row=row_number, column=3).value
    if month == month_start:     # если равна, то запускается следующий цикл
        for i in range(row_number+1, sheet.max_row+1):
            document_if = sheet.cell(row=i, column=7).value
            receiving_date = sheet.cell(row=i, column=8).value

            if document_if == document and receiving_date > ttt and receiving_date < rrr:
                document_amount +=1

            if month == month_stop: # если цикл доходит до конца периода, то цикл обрывается
                break
    if month == month_stop: # конец периода выборки = обрыв цикла
        break

print( "Количество договоров по майским сделкам в июне", str(document_amount) )

# сохраняю все полученные данные в книге data4.xlsx в той же директории
sheet['J2'] = "Количество договоров по майским сделкам в июне " + str(document_amount)
sheet.column_dimensions['J'].width = 45
book.save('data4.xlsx')