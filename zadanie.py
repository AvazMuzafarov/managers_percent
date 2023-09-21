# Аваз Музафаров, 21.09.2021
# За каждую заключенную сделку менеджер получает бонус, который рассчитывается
# следующим образом.
# 1) За новые сделки менеджер получает 7 % от суммы, при условии, что статус
# оплаты «ОПЛАЧЕНО», а также имеется оригинал подписанного договора с
# клиентом (в рассматриваемом месяце).
# 2) За текущие сделки менеджер получает 5 % от суммы, если она больше 10 тыс.,
# и 3 % от суммы, если меньше. При этом статус оплаты может быть любым,
# кроме «ПРОСРОЧЕНО», а также необходимо наличие оригинала подписанного
# договора с клиентом (в рассматриваемом месяце).
# Бонусы по сделкам, оригиналы для которых приходят позже рассматриваемого
# месяца, считаются остатком на следующий период, который выплачивается по мере
# прихода оригиналов. Вычислите остаток каждого из менеджеров на 01.07.2021.

# загрузка модуля openpyxl для работы с файлами Excel и модуля datetime
import openpyxl
import datetime

# создание книги и листа
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active

# задаю все необходимые переменные
month_start = 'Июнь 2021'                                   # начало периода выборки
month_stop = 'Июль 2021'                                    # конец периода выборки
date_start = datetime.datetime(2021, 5, 31, 23, 59, 11 )    # начало периода выборки договоров (datetime)
date_stop = datetime.datetime(2021, 7, 1, 1, 11, 11 )       # конец периода выборки  договоров (datetime)
need_doc = 'оригинал'                                       # необходимый формат договора
type_new = 'новая'                                          # тип сделки новая
type_current = 'текущая'                                    # тип сделки текущая
need_status = 'ОПЛАЧЕНО'                                    # необходимый статус оплаты
bad_status = 'ПРОСРОЧЕНО'                                   # невалидный статус оплаты
managers = {}                                               # словарь хранящий всю информацию о менеджерах и их средствах

# главный цикл, проверяет строки на наличие month_start
for row_number in range(2, sheet.max_row+1):
    month = sheet.cell(row=row_number, column=3).value

    if month == month_start:     # если начало периода определено, то запускается 2й цикл
        for i in range(row_number+1, sheet.max_row+1):  # этот цикл проверяет каждую строку в указанном месяце

            sum = sheet.cell(row=i, column=2).value
            status = sheet.cell(row=i, column=3).value
            sale = sheet.cell(row=i, column=4).value
            new_current = sheet.cell(row=i, column=5).value
            document = sheet.cell(row=i, column=7).value
            receiving_date = sheet.cell(row=i, column=8).value

            managers.setdefault(sale, {'bonus_live' : 0, 'bonus_os' : 0})    # задается ключ-менеджер для словаря managers
                                                                             # и задаются 2 ключа каждому менеджеру
                                                                             # bonus_live - бонусы за month_start
                                                                             # bonus_os - остатки на конец month_start (на 01.07.2021)

            if new_current == type_new and status == need_status and document == need_doc: # проверка на тип "новая"
                if receiving_date > date_start and receiving_date < date_stop:  # проверка на дату получения договора
                    bonus1 = sum * 0.07
                    managers[sale]['bonus_live'] += bonus1     # дата актуальна = бонус зачисляется менеджеру в этом месяце
                if receiving_date > date_stop:
                    bonus1 = sum * 0.07
                    managers[sale]['bonus_os'] += bonus1     # дата не актуальна = бонус уходит в остаток
            if new_current == type_current and status != bad_status and document == need_doc: # проверка на тип "текущая"
                if receiving_date > date_start and receiving_date < date_stop:                # по сути здесь тоже самое
                    if sum >= 10000:
                        bonus2 = sum * 0.05
                        managers[sale]['bonus_live'] += bonus2
                    if sum < 10000:
                        bonus3 = sum * 0.03
                        managers[sale]['bonus_live'] += bonus3
                if receiving_date > date_stop:
                    if sum >= 10000:
                        bonus2 = sum * 0.05
                        managers[sale]['bonus_os'] += bonus2
                    if sum < 10000:
                        bonus3 = sum * 0.03
                        managers[sale]['bonus_os'] += bonus3


            if month == month_stop: # если цикл доходит до конца периода, то цикл обрывается
                break
    if month == month_stop: # конец периода выборки = обрыв цикла
        break

print(managers)

# сохраняю все полученные данные в книге data_zadanie.xlsx в той же директории
sheet['J3'] = "Менеджер"
sheet['K3'] = "Бонус за месяц"
sheet['L3'] = "Остаток"

managers.pop(None, 'Мертвой души нет!')  # проверка на мертвую душу
managers_list = []                       # список менеджеров
managers_list_len = len(managers_list)   # количество менеджеров
row_row = 3                              # переменная для указания строки вывода
for name in managers:                    # цикл, который заполняет все ячейки столбцов "менеджер", "Бонус за месяц" и "Остаток"
    row_row +=1
    sheet['J' + str(row_row)] = name
    sheet['K' + str(row_row)] = str(managers[name]['bonus_live'])
    sheet['L' + str(row_row)] = str(managers[name]['bonus_os'])

book.save('data_zadanie.xlsx')    # сохраняю все в книге data_zadanie.xlsx