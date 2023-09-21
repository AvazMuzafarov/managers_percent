# Аваз Музафаров, 21.09.2023
# 1) Вычислите общую выручку за июль 2021 по тем сделкам, приход денежных средств которых не просрочен.
# 2) Как изменилась выручка компании за рассматриваемый период? Проилюстрируйте графиком. (рассматриваемый период = июль 2021)
# Я решил ответить на оба вопроса сразу, так как 2-ой связан с первым.

# загрузка модуля openpyxl для работы с файлами Excel
import openpyxl

# создание книги и листа
book = openpyxl.load_workbook('data.xlsx')
sheet = book.active

# задаю все необходимые переменные
month_start = 'Июль 2021'    # месяц, начало периода выборки
month_stop = 'Август 2021'   # месяц, конец периода
status_paid = 'ОПЛАЧЕНО'     # статусы сделок, кторые войдут в выборку
paid_sum = 0                 # переменная - общая сумма сделок за период
paid_list_dif = 1            # начало списка сделок выборки (нужно для графика)
paid_list_len = 0            # конец списка сделок выборки (нужно для графика)


# собственно сам цикл, который будет перебирать все строки листа и проверять, не равна ли переменная month началу отсчета
for row_number in range(2, sheet.max_row+1):
    month = sheet.cell(row=row_number, column=3).value
    if month == month_start:     # если равна, то запускается следующий цикл
        for i in range(row_number+1, sheet.max_row+1):  # этот цикл проверяет статус сделки в выбрнном периоде (от month_start до month_stop)
            status = sheet.cell(row=i, column=3).value
            sum = sheet.cell(row=i, column=2).value

            if status == status_paid: # если статус сделки 'ОПЛАЧЕНО', то сумма этой сделки сохраняется в paid_sum
                paid_sum += sum

                # а также, эта сумма сохранится в виде элемента списка в столбце L книги data1.xlsx (нужно для графика)
                paid_list_dif +=1
                sheet['L' + str(paid_list_dif)] = sum
                paid_list_len +=1

            if status == month_stop: # если цикл доходит до конца периода, то цикл обрывается
                break
    if month == month_stop: # конец периода выборки = обрыв цикла
        break

# создание графика, для наглядности изменения выручки за период
ref_obj = openpyxl.chart.Reference(sheet, min_col=12, min_row=2, max_col=12, max_row =paid_list_len )
series_obj = openpyxl.chart.Series(ref_obj, title = 'Сумма выручки')
chart_obj = openpyxl.chart.LineChart()
chart_obj.title = 'Выручка за ' + month_start
chart_obj.append(series_obj)
sheet.add_chart(chart_obj, 'J5')

# сохраняю все полученные данные в книге data1.xlsx в той же директории
sheet['J2'] = 'Выручка за ' + month_start
sheet.column_dimensions['J'].width = 24
sheet['J3'] = paid_sum
book.save('data1.xlsx')

# Ответ 1: это число (859896.469), которое находится в ячейке J3 книги data1.xlsx.
# Ответ 2: график создан и расположен в ячейке J5  книги data1.xlsx.